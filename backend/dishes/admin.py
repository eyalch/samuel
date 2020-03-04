import datetime
from collections import defaultdict
from tempfile import NamedTemporaryFile

from babel.dates import format_date
from django import forms
from django.conf import settings
from django.contrib import admin, messages
from django.core.mail import EmailMultiAlternatives, send_mass_mail
from django.db.models import Count, F
from django.http import HttpResponse, HttpResponseRedirect
from django.template.loader import get_template
from django.urls import path
from django.utils import timezone
from django.utils.html import mark_safe
from openpyxl import Workbook
from rangefilter.filter import DateRangeFilter

from .errors import TimeIsUpError
from .models import Dish, Order, ScheduledDish, TodayOrder
from .views import check_if_time_is_up_for_today

plaintext_template = get_template("email/daily_dishes.txt")
html_template = get_template("email/daily_dishes.html")

XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def auto_adjust_columns(ws):
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 1


class AddScheduledDishInline(admin.TabularInline):
    model = ScheduledDish
    max_num = 1
    can_delete = False
    verbose_name_plural = "new scheduled dish"

    def has_view_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        return False


class DishesEmailModelAdminMixin:
    change_list_template = "admin/dishes_changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        my_urls = [
            path("send_email/", self.send_email),
            path("send_email_test/", self.send_email, {"test": True}),
            path("send_dishes_ready/", self.send_dishes_ready),
        ]
        return my_urls + urls

    def send_email(self, request, test=False):
        """
        Send an email with today's dishes. If the time is up for today, then send
        tomorrow's dishes. The email will be sent only if there are scheduled dishes.
        """
        date = timezone.now().date()
        try:
            check_if_time_is_up_for_today()
        except TimeIsUpError:
            date += datetime.timedelta(days=1)

        scheduled_dishes = ScheduledDish.objects.filter(date=date)

        if len(scheduled_dishes) == 0:
            self.message_user(
                request, "No scheduled dishes. Not sending an email.", messages.WARNING
            )
            return HttpResponseRedirect("../")

        dishes = [scheduled_dish.dish for scheduled_dish in scheduled_dishes]
        context = {
            "dishes": dishes,
            "base_url": settings.BASE_URL,
            "weekday": format_date(date, "EEE", locale="he"),
            "date": format_date(date, "d בMMMM", locale="he"),
            "test": test,
        }

        EmailMultiAlternatives(
            to=settings.EMAIL_TEST_RECIPIENTS if test else settings.EMAIL_RECIPIENTS,
            bcc=settings.EMAIL_TEST_RECIPIENTS if not test else None,
            subject="סמואל",
            body=plaintext_template.render(context),
            alternatives=[(html_template.render(context), "text/html")],
        ).send()

        success_message = (
            "Successfully sent an email."
            if not test
            else "Successfully sent a test email."
        )

        self.message_user(request, success_message, messages.SUCCESS)

        return HttpResponseRedirect("../")

    def send_dishes_ready(self, request):
        today = timezone.now().date()
        dishes_scheduled_for_today = ScheduledDish.objects.filter(date=today)

        datatuple = []

        for scheduled_dish in dishes_scheduled_for_today:
            # Get all ordering users' email addresses for the ScheduledDish
            orders = Order.objects.select_related("user").filter(
                scheduled_dish=scheduled_dish
            )
            email_addresses = [order.user.email for order in orders]

            subject = f'המנה "{scheduled_dish.dish}" מוכנה!'
            body = "בתיאבון!"

            for email_address in email_addresses:
                datatuple.append((subject, body, None, (email_address,)))

        messages_sent = send_mass_mail(datatuple)

        if messages_sent > 0:
            self.message_user(request, "Messages sent successfully.", messages.SUCCESS)
        else:
            self.message_user(request, "No messages sent.", messages.WARNING)

        return HttpResponseRedirect("../")


class DishAdminForm(forms.ModelForm):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)

        # Make the "name" and "description" fields RTL
        self.fields["name"].widget.attrs.update({"style": "direction: rtl"})
        self.fields["description"].widget.attrs.update(
            {"style": "direction: rtl; width: auto", "rows": 5}
        )


@admin.register(Dish)
class DishAdmin(DishesEmailModelAdminMixin, admin.ModelAdmin):
    list_display = ("name", "description", "image_preview")
    actions = ["schedule_for_today", "schedule_for_tomorrow"]
    form = DishAdminForm
    inlines = [AddScheduledDishInline]
    actions_on_bottom = True
    radio_fields = {"dish_type": admin.VERTICAL}
    search_fields = ["name"]
    ordering = ["name"]
    fields = ("name", "description", ("image", "image_preview"), "dish_type")
    readonly_fields = ["image_preview"]

    def image_preview(self, obj):
        if obj.image:
            return mark_safe(f'<img src="{obj.image.url}" height="120" />')

    def schedule_for(self, dishes, date, request, pretty_day):
        """Schedule the given dishes for the given date"""

        scheduled_dishes_created = 0

        for dish in dishes:
            _, created = ScheduledDish.objects.get_or_create(dish=dish, date=date)
            if created:
                scheduled_dishes_created += 1

        if scheduled_dishes_created == 1:
            message_bit = "1 dish was"
        else:
            message_bit = f"{scheduled_dishes_created} dishes were"
        self.message_user(request, f"{message_bit} scheduled for {pretty_day}.")

    def schedule_for_today(self, request, queryset):
        today = timezone.now().date()
        self.schedule_for(queryset, today, request, "today")

    schedule_for_today.short_description = "Schedule for today"

    def schedule_for_tomorrow(self, request, queryset):
        tomorrow = timezone.now().date() + datetime.timedelta(days=1)
        self.schedule_for(queryset, tomorrow, request, "tomorrow")

    schedule_for_tomorrow.short_description = "Schedule for tomorrow"


@admin.register(ScheduledDish)
class ScheduledDishAdmin(DishesEmailModelAdminMixin, admin.ModelAdmin):
    list_display = ("__str__", "orders_left")
    ordering = ("-date",)
    list_filter = ("date", "dish")
    date_hierarchy = "date"
    autocomplete_fields = ["dish"]


def add_orders_header_row(ws):
    ws.cell(row=1, column=1, value="Name")
    ws.cell(row=1, column=2, value="Email")

    # Add a column for each dish type
    for index, dish_type_label in enumerate(Dish.DishType.labels, 3):
        ws.cell(row=1, column=index, value=dish_type_label)

    # Freeze the header row
    ws.freeze_panes = "A2"


def add_orders_user_row(ws, user, user_orders, index):
    # Iterate the user's orders and count the orders of each dish type
    orders_count_by_dish_type = defaultdict(int)
    for order in user_orders:
        orders_count_by_dish_type[order.scheduled_dish.dish.dish_type] += 1

    ws.cell(row=index, column=1, value=str(user))
    ws.cell(row=index, column=2, value=str(user.email))

    # Iterate the dish types and put the count of each dish type into it's own column
    for i, dish_type in enumerate(Dish.DishType.values, 3):
        count = orders_count_by_dish_type.get(dish_type, 0)
        ws.cell(row=index, column=i, value=count)


@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ("scheduled_dish", "user", "get_dish_type", "created_at")
    list_filter = (
        ("scheduled_dish__date", DateRangeFilter),
        "scheduled_dish__date",
        "scheduled_dish__dish",
    )
    list_display_links = None
    ordering = ("-scheduled_dish__date",)
    date_hierarchy = "scheduled_dish__date"
    actions = ("export_selected_orders",)

    def get_dish_type(self, obj):
        return obj.scheduled_dish.dish.dish_type

    get_dish_type.short_description = "Dish type"
    get_dish_type.admin_order_field = "scheduled_dish__dish__dish_type"

    def has_change_permission(self, request, obj=None):
        return False

    def export_selected_orders(self, request, queryset):
        # Group the orders by user and then by domain
        orders_by_user_by_domain = defaultdict(lambda: defaultdict(list))
        for order in queryset:
            orders_by_user_by_domain[order.user.domain][order.user].append(order)

        wb = Workbook()
        main_ws = wb.active

        # Prepare the main orders sheet (w/ all users)
        main_ws.title = "All"
        add_orders_header_row(main_ws)

        all_users_index = 2

        for domain, orders_by_user in orders_by_user_by_domain.items():
            # Prepare the domain orders sheet (w/ users only from that domain)
            domain_ws = wb.create_sheet(domain)
            add_orders_header_row(domain_ws)

            for index, (user, user_orders) in enumerate(orders_by_user.items(), 2):
                # Add a row of the user to the domain sheet
                add_orders_user_row(domain_ws, user, user_orders, index)

                # Add a row of the user to the main sheet
                add_orders_user_row(main_ws, user, user_orders, all_users_index)
                all_users_index += 1

            auto_adjust_columns(domain_ws)

        auto_adjust_columns(main_ws)

        # Save to a temporary file
        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(stream, content_type=XLSX_MIME_TYPE)
        response["Content-Disposition"] = "attachment; filename=orders.xlsx"
        return response

    export_selected_orders.short_description = "Export selected orders"


@admin.register(TodayOrder)
class OrdersForToday(admin.ModelAdmin):
    list_display = ("user", "get_dish", "created_at")
    list_display_links = None
    ordering = ("-created_at",)
    change_list_template = "admin/today_orders_changelist.html"

    def get_dish(self, obj):
        return obj.scheduled_dish.dish

    get_dish.short_description = "Dish"

    def has_add_permission(self, request, obj=None):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}

        orders_count_by_dish = (
            TodayOrder.objects.values(
                dish_id=F("scheduled_dish__dish"), name=F("scheduled_dish__dish__name")
            )
            .annotate(orders_count=Count("dish_id"))
            .values("name", "orders_count")
        )
        extra_context["orders_count_by_dish"] = list(orders_count_by_dish)

        return super().changelist_view(request, extra_context=extra_context)

    def get_urls(self):
        return [path("export/", self.export_today_orders)] + super().get_urls()

    def export_today_orders(self, request):
        wb = Workbook()
        ws = wb.active

        for index, order in enumerate(TodayOrder.objects.all(), 1):
            ws.cell(row=index, column=1, value=str(order.user))
            ws.cell(row=index, column=2, value=str(order.scheduled_dish.dish))

        auto_adjust_columns(ws)

        with NamedTemporaryFile() as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            stream = tmp.read()

        response = HttpResponse(stream, content_type=XLSX_MIME_TYPE)
        response["Content-Disposition"] = "attachment; filename=today_orders.xlsx"
        return response
